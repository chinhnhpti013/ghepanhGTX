import openpyxl, re, os, io, sys
from PIL import Image, ImageDraw, ImageFont, ImageOps
import numpy as np

# Fix encoding for Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

DPI = 150
PORTRAIT_W, PORTRAIT_H   = 1240, 1753
LANDSCAPE_W, LANDSCAPE_H = 1753, 1240
MARGIN    = 30
LOGO_PX   = 107
HEADER_H  = 130
FOOTER_H  = 50
CAPTION_H = 38
BLUE  = (0, 112, 192)
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
GRAY  = (200, 200, 200)
FONT_DIR  = r"C:\Windows\Fonts"

def pt2px(pt):
    return int(pt * DPI / 72)

def get_font(bold=False, italic=False, pt=12):
    px = pt2px(pt)
    if bold:
        paths = [os.path.join(FONT_DIR, "arialbd.ttf")]
    elif italic:
        paths = [os.path.join(FONT_DIR, "ariali.ttf")]
    else:
        paths = [os.path.join(FONT_DIR, "arial.ttf")]
    for p in paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, px)
            except:
                pass
    return ImageFont.load_default()

def process_logo(logo_path, size_px):
    logo = Image.open(logo_path).convert('RGBA')
    data = np.array(logo)
    r, g, b, a = data[:,:,0], data[:,:,1], data[:,:,2], data[:,:,3]
    is_dark = (r < 80) & (g < 80) & (b < 80)
    data[is_dark] = [0, 0, 0, 0]
    is_blue = (r < 60) & (g < 60) & (b > 100) & (~is_dark)
    data[is_blue] = [255, 255, 255, 255]
    return Image.fromarray(data).resize((size_px, size_px), Image.LANCZOS)

def read_gtx_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    bks = ""
    # Collect all cell values from first 8 rows
    all_cells = []
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        all_cells.extend([c for c in row if c])
    # Find plate number pattern anywhere
    for cell in all_cells:
        if isinstance(cell, str):
            m = re.search(r'\d{2}[A-Z]-?\d{3,5}\.?\d*', cell)
            if m:
                bks = m.group(0)
                break
    items = {}
    for row in ws.iter_rows(values_only=True):
        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None
        if c0 is None or not (c1 and isinstance(c1, str)):
            continue
        # STT có thể là số nguyên (3) hoặc thập phân (1.1, 1.2). Chuẩn hoá thành
        # chuỗi key khớp với phần đầu tên file: "3", "1.1", "1.2"...
        if isinstance(c0, int) and c0 > 0:
            key = str(c0)
        elif isinstance(c0, float) and c0 > 0:
            key = str(int(c0)) if c0.is_integer() else str(c0)
        elif isinstance(c0, str) and re.match(r'^\d+(\.\d+)?$', c0.strip()):
            key = c0.strip()
        else:
            continue
        items[key] = str(c1).strip()
    return bks, items

def classify_images(image_paths, items):
    """Phân loại ảnh theo số nhóm đầu tên file (ví dụ: 1.1, 1.2, 2.1, 3.jpg).
    Trả về dict {nhóm: [(path, caption, sort_key), ...]} sắp xếp theo tên file.
    """
    classified = {}
    for path in image_paths:
        fname = os.path.basename(path)
        name_no_ext = os.path.splitext(fname)[0]
        # Khớp key STT dạng thập phân ("1.2") lẫn số nguyên ("3", "4_001_pdf")
        m = re.match(r'^(\d+(?:\.\d+)?)', name_no_ext)
        if m:
            key = m.group(1)            # "1.1", "2.2", "3", "4"...
            nhom = int(float(key))      # nhóm số nguyên để chọn layout
            ten = items.get(key) or items.get(str(nhom))
            if not ten:
                # Thiếu STT trong Excel: dùng tên của mục gần nhất phía trước cùng nhóm
                prev_keys = sorted(
                    (k for k in items if k.startswith(f'{nhom}.') and float(k) < float(key)),
                    key=float
                )
                ten = items[prev_keys[-1]] if prev_keys else f'Hạng mục {key}'
            caption = f"{key}. {ten}"
            if nhom not in classified:
                classified[nhom] = []
            classified[nhom].append((path, caption, name_no_ext))
    for nhom in classified:
        classified[nhom].sort(key=lambda x: x[2])
    return classified

def wrap_text(draw, text, font, max_w):
    words = text.split()
    lines = []
    cur = ""
    for w in words:
        test = (cur + " " + w).strip()
        bb = draw.textbbox((0, 0), test, font=font)
        if bb[2] - bb[0] <= max_w:
            cur = test
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines

def draw_header(page, draw, W, HEADER_H, MARGIN, logo_px, logo_img, bks, gdv, ngay=''):
    draw.rectangle([0, 0, W, HEADER_H], fill=BLUE)
    logo_y = (HEADER_H - logo_px) // 2
    page.paste(logo_img, (MARGIN, logo_y), logo_img)
    text_x = MARGIN + logo_px + 10
    text_w = W - text_x - MARGIN
    font_title = get_font(bold=True, pt=16)
    font_gdv   = get_font(bold=False, pt=12)
    title_lines = wrap_text(draw, f"Giam dinh Giay to xe o to bien kiem soat: {bks}", font_title, text_w)
    # Use actual Vietnamese text
    title_lines = wrap_text(draw, f"Giám định Giấy tờ xe ô tô biển kiểm soát: {bks}", font_title, text_w)
    tlh = pt2px(16) + 4
    total_th = len(title_lines) * tlh
    gdv_h = pt2px(12) + 4 if gdv else 0
    cur_y = (HEADER_H - total_th - gdv_h) // 2
    for line in title_lines:
        bb = draw.textbbox((0, 0), line, font=font_title)
        x = text_x + (text_w - (bb[2] - bb[0])) // 2
        draw.text((x, cur_y), line, font=font_title, fill=WHITE)
        cur_y += tlh
    if gdv:
        gdv_text = f"Giám định viên: {gdv}"
        if ngay:
            gdv_text += f" - Ngày: {ngay}"
        bb = draw.textbbox((0, 0), gdv_text, font=font_gdv)
        x = text_x + (text_w - (bb[2] - bb[0])) // 2
        draw.text((x, cur_y), gdv_text, font=font_gdv, fill=WHITE)

def draw_footer(page, draw, W, H, FOOTER_H, MARGIN, page_num, total_pages):
    fy = H - FOOTER_H
    draw.rectangle([0, fy, W, H], fill=WHITE)
    draw.line([0, fy, W, fy], fill=GRAY, width=1)
    font_f  = get_font(bold=False, pt=10)
    font_fi = get_font(italic=True, pt=10)
    ty = fy + (FOOTER_H - pt2px(10)) // 2
    draw.text((MARGIN, ty), "Phòng Giám định và Cứu hộ tại Quảng Ninh", font=font_f, fill=(128, 128, 128))
    pt_text = f"Trang {page_num}/{total_pages}"
    bb = draw.textbbox((0, 0), pt_text, font=font_fi)
    draw.text((W - MARGIN - (bb[2] - bb[0]), ty), pt_text, font=font_fi, fill=BLACK)

def draw_caption(page, draw, caption, cx, cy, cell_w, cell_h, CAPTION_H):
    cap_y = cy + cell_h - CAPTION_H
    draw.rectangle([cx, cap_y, cx + cell_w, cy + cell_h], fill=WHITE)
    draw.line([cx, cap_y, cx + cell_w, cap_y], fill=GRAY, width=1)
    font_cap = get_font(bold=False, pt=11)
    bb = draw.textbbox((0, 0), caption, font=font_cap)
    draw.text(
        (cx + (cell_w - (bb[2] - bb[0])) // 2, cap_y + (CAPTION_H - (bb[3] - bb[1])) // 2),
        caption, font=font_cap, fill=BLACK
    )

def paste_image_in_cell(page, img_path, cx, cy, cell_w, img_area_h):
    try:
        img = Image.open(img_path)
        img = ImageOps.exif_transpose(img).convert('RGB')
        iw, ih = img.size
        scale = min(cell_w / iw, img_area_h / ih)
        nw, nh = int(iw * scale), int(ih * scale)
        img_r = img.resize((nw, nh), Image.LANCZOS)
        page.paste(img_r, (cx + (cell_w - nw) // 2, cy + (img_area_h - nh) // 2))
    except Exception as e:
        print(f"Loi anh {img_path}: {e}")

def create_landscape_page(imgs, page_num, total, hinfo, logo_img):
    W, H = LANDSCAPE_W, LANDSCAPE_H
    COLS, ROWS = 2, 2
    page = Image.new('RGB', (W, H), WHITE)
    draw = ImageDraw.Draw(page)
    draw_header(page, draw, W, HEADER_H, MARGIN, LOGO_PX, logo_img, hinfo['bks'], hinfo.get('gdv', ''), hinfo.get('ngay', ''))
    draw_footer(page, draw, W, H, FOOTER_H, MARGIN, page_num, total)
    draw.rectangle([0, 0, W - 1, H - 1], outline=GRAY, width=1)
    cw = (W - 2 * MARGIN) // COLS
    ch = (H - HEADER_H - FOOTER_H) // ROWS
    iah = ch - CAPTION_H
    for idx, (ip, cap) in enumerate(imgs[:COLS * ROWS]):
        col, row = idx % COLS, idx // COLS
        cx = MARGIN + col * cw
        cy = HEADER_H + row * ch
        draw.rectangle([cx, cy, cx + cw, cy + ch], outline=GRAY, width=1)
        paste_image_in_cell(page, ip, cx, cy, cw, iah)
        draw_caption(page, draw, cap, cx, cy, cw, ch, CAPTION_H)
    for idx in range(len(imgs), COLS * ROWS):
        col, row = idx % COLS, idx // COLS
        cx = MARGIN + col * cw
        cy = HEADER_H + row * ch
        draw.rectangle([cx, cy, cx + cw, cy + ch], outline=GRAY, width=1)
    return page

def create_portrait_page_2(imgs, page_num, total, hinfo, logo_img):
    """A4 dọc, 2 ảnh xếp theo 1 cột × 2 hàng."""
    W, H = PORTRAIT_W, PORTRAIT_H
    ROWS = 2
    page = Image.new('RGB', (W, H), WHITE)
    draw = ImageDraw.Draw(page)
    draw_header(page, draw, W, HEADER_H, MARGIN, LOGO_PX, logo_img, hinfo['bks'], hinfo.get('gdv', ''), hinfo.get('ngay', ''))
    draw_footer(page, draw, W, H, FOOTER_H, MARGIN, page_num, total)
    draw.rectangle([0, 0, W - 1, H - 1], outline=GRAY, width=1)
    cw = W - 2 * MARGIN
    ch = (H - HEADER_H - FOOTER_H) // ROWS
    iah = ch - CAPTION_H
    for idx, (ip, cap) in enumerate(imgs[:ROWS]):
        cx = MARGIN
        cy = HEADER_H + idx * ch
        draw.rectangle([cx, cy, cx + cw, cy + ch], outline=GRAY, width=1)
        paste_image_in_cell(page, ip, cx, cy, cw, iah)
        draw_caption(page, draw, cap, cx, cy, cw, ch, CAPTION_H)
    for idx in range(len(imgs), ROWS):
        cx = MARGIN
        cy = HEADER_H + idx * ch
        draw.rectangle([cx, cy, cx + cw, cy + ch], outline=GRAY, width=1)
    return page

def create_portrait_page_1(ip, cap, page_num, total, hinfo, logo_img):
    W, H = PORTRAIT_W, PORTRAIT_H
    page = Image.new('RGB', (W, H), WHITE)
    draw = ImageDraw.Draw(page)
    draw_header(page, draw, W, HEADER_H, MARGIN, LOGO_PX, logo_img, hinfo['bks'], hinfo.get('gdv', ''), hinfo.get('ngay', ''))
    draw_footer(page, draw, W, H, FOOTER_H, MARGIN, page_num, total)
    draw.rectangle([0, 0, W - 1, H - 1], outline=GRAY, width=1)
    cx = MARGIN; cy = HEADER_H
    cw = W - 2 * MARGIN; ch = H - HEADER_H - FOOTER_H
    iah = ch - CAPTION_H
    draw.rectangle([cx, cy, cx + cw, cy + ch], outline=GRAY, width=1)
    paste_image_in_cell(page, ip, cx, cy, cw, iah)
    draw_caption(page, draw, cap, cx, cy, cw, ch, CAPTION_H)
    return page

def pdf_to_images(pdf_path, stt, tmp_dir):
    """Chuyển PDF sang list ảnh JPG tạm, trả về list đường dẫn."""
    import fitz
    doc = fitz.open(pdf_path)
    paths = []
    for i, page in enumerate(doc):
        mat = fitz.Matrix(150/72, 150/72)
        pix = page.get_pixmap(matrix=mat)
        out = os.path.join(tmp_dir, f"{stt}_{i+1:03d}_pdf.jpg")
        pix.save(out)
        paths.append(out)
    doc.close()
    return paths

def main(excel_path, logo_path, image_dir, output_path, gdv='', ngay=''):
    import tempfile
    bks, items = read_gtx_excel(excel_path)
    # Update output filename with actual BKS
    if bks and 'GiamDinh_GiayToXe.pdf' in output_path:
        output_path = output_path.replace('GiamDinh_GiayToXe.pdf', f'GiamDinh_GiayToXe_{bks}.pdf')
    print(f"Bien so xe: {bks}")
    print(f"Danh muc: {items}")

    tmp_dir = tempfile.mkdtemp()

    # Chuyển PDF thành ảnh trước
    pdf_image_paths = []
    for f in sorted(os.listdir(image_dir)):
        if f.lower().endswith('.pdf'):
            m = re.match(r'^(\d+)', f)
            if m:
                nhom = int(m.group(1))
                converted = pdf_to_images(os.path.join(image_dir, f), nhom, tmp_dir)
                pdf_image_paths.extend(converted)
                print(f"Chuyển PDF {f} -> {len(converted)} ảnh")

    image_paths = [
        os.path.join(image_dir, f)
        for f in os.listdir(image_dir)
        if f.lower().endswith(('.jpg', '.jpeg', '.png')) and 'logo' not in f.lower()
    ] + pdf_image_paths
    classified = classify_images(image_paths, items)
    print(f"Phan loai anh: { {k: [x[2] for x in v] for k, v in classified.items()} }")

    logo_img = process_logo(logo_path, LOGO_PX)
    hinfo = {'bks': bks, 'gdv': gdv, 'ngay': ngay}
    pages_info = []

    # Nhóm 1 — A4 ngang, tối đa 4 ảnh/trang (2 cột × 2 hàng)
    if 1 in classified:
        for i in range(0, len(classified[1]), 4):
            pages_info.append(('landscape', classified[1][i:i+4]))

    # Nhóm 2 — A4 dọc, tối đa 2 ảnh/trang (1 cột × 2 hàng)
    if 2 in classified:
        for i in range(0, len(classified[2]), 2):
            pages_info.append(('portrait2', classified[2][i:i+2]))

    # Nhóm 3 — A4 dọc, tối đa 2 ảnh/trang (1 cột × 2 hàng)
    if 3 in classified:
        for i in range(0, len(classified[3]), 2):
            pages_info.append(('portrait2', classified[3][i:i+2]))

    # Nhóm khác (4, 5, 6...) — A4 dọc, 1 ảnh/trang
    for nhom in sorted(k for k in classified if k >= 4):
        for item in classified[nhom]:
            pages_info.append(('portrait1', [item]))

    total = len(pages_info)
    print(f"Tổng số trang: {total}")

    if total == 0:
        raise ValueError(
            "Không có ảnh nào được phân loại. Kiểm tra tên file phải bắt đầu bằng "
            "số nhóm (ví dụ 1.1.jpg, 2.1.jpg, 4.pdf)."
        )

    from reportlab.pdfgen import canvas as rlc
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    c = rlc.Canvas(output_path)

    for p_idx, (ptype, data) in enumerate(pages_info, 1):
        if ptype == 'landscape':
            pg = create_landscape_page([(d[0], d[1]) for d in data], p_idx, total, hinfo, logo_img)
            pw, ph = 29.7 * cm, 21 * cm
        elif ptype == 'portrait2':
            pg = create_portrait_page_2([(d[0], d[1]) for d in data], p_idx, total, hinfo, logo_img)
            pw, ph = 21 * cm, 29.7 * cm
        else:
            d = data[0]
            pg = create_portrait_page_1(d[0], d[1], p_idx, total, hinfo, logo_img)
            pw, ph = 21 * cm, 29.7 * cm

        c.setPageSize((pw, ph))
        buf = io.BytesIO()
        pg.save(buf, format='JPEG', quality=92)
        buf.seek(0)
        c.drawImage(ImageReader(buf), 0, 0, pw, ph)
        c.showPage()

    c.save()
    print(f"Xuat PDF: {output_path}")
    return output_path

if __name__ == '__main__':
    BASE = r"d:\MCP\Claude Code\anh-giay-to-xe"
    main(
        excel_path=os.path.join(BASE, "input", "GTX.xlsx"),
        logo_path=os.path.join(BASE, "assets", "logo-ptisos.png"),
        image_dir=os.path.join(BASE, "input"),
        output_path=os.path.join(BASE, "output", "GiamDinh_GiayToXe.pdf"),
        gdv="Nguyễn Văn Hướng",
        ngay="14/07/2026"
    )

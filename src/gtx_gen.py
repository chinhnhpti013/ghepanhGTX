import openpyxl, re, os, io
from PIL import Image, ImageDraw, ImageFont
import numpy as np

DPI=150; PORTRAIT_W,PORTRAIT_H=1240,1753; LANDSCAPE_W,LANDSCAPE_H=1753,1240
MARGIN=30; LOGO_PX=89; HEADER_H=130; FOOTER_H=50; CAPTION_H=38
NAVY=(0,32,96); WHITE=(255,255,255); BLACK=(0,0,0); GRAY=(200,200,200)

BASE_DIR   = r"d:\MCP\Claude Code\anh-giay-to-xe"
EXCEL_PATH = os.path.join(BASE_DIR, "input", "GTX.xlsx")
LOGO_PATH  = os.path.join(BASE_DIR, "assets", "logo-ptisos-fixed.png")
IMAGE_DIR  = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
GDV        = "Nguyễn Thế Duy"
BKS_FALLBACK = "14K-239.42"

def pt2px(pt): return int(pt*DPI/72)

def get_font(bold=False, italic=False, pt=12):
    px = pt2px(pt)
    win_paths_bold   = [r"C:\Windows\Fonts\arialbd.ttf", r"C:\Windows\Fonts\calibrib.ttf"]
    win_paths_italic = [r"C:\Windows\Fonts\ariali.ttf",  r"C:\Windows\Fonts\calibrii.ttf"]
    win_paths_reg    = [r"C:\Windows\Fonts\arial.ttf",   r"C:\Windows\Fonts\calibri.ttf"]
    lin_paths_bold   = ["/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"]
    lin_paths_italic = ["/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
                        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf"]
    lin_paths_reg    = ["/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
    paths = (win_paths_bold + lin_paths_bold if bold else
             win_paths_italic + lin_paths_italic if italic else
             win_paths_reg + lin_paths_reg)
    for p in paths:
        if os.path.exists(p):
            try: return ImageFont.truetype(p, px)
            except: pass
    return ImageFont.load_default()

def process_logo(logo_path, size_px):
    logo = Image.open(logo_path).convert('RGBA')
    data = np.array(logo)
    r,g,b,a = data[:,:,0],data[:,:,1],data[:,:,2],data[:,:,3]
    is_dark = (r<80)&(g<80)&(b<80)
    data[is_dark] = [0,0,0,0]
    is_navy = (r<60)&(g<60)&(b>100)&(~is_dark)
    data[is_navy] = [255,255,255,255]
    return Image.fromarray(data).resize((size_px,size_px), Image.LANCZOS)

def read_gtx_excel(path, fallback_bks=""):
    wb = openpyxl.load_workbook(path); ws = wb.active; bks=""
    for row in ws.iter_rows(min_row=1,max_row=5,values_only=True):
        for cell in row:
            if cell and isinstance(cell,str) and 'Giấy tờ xe' in cell:
                parts=cell.strip().split()
                candidate = parts[-1] if parts else ""
                if candidate and candidate != 'xe': bks = candidate
                break
        if bks: break
    if not bks: bks = fallback_bks
    items={}
    for row in ws.iter_rows(values_only=True):
        c0=row[0]; c1=row[1] if len(row)>1 else None
        if isinstance(c0,int) and c0>0 and c1 and isinstance(c1,str):
            items[c0]=str(c1).strip()
    return bks,items

def classify_images(image_paths,items):
    classified={}
    for path in image_paths:
        fname=os.path.basename(path); name_no_ext=os.path.splitext(fname)[0]
        m=re.match(r'^(\d+)',name_no_ext)
        if m:
            stt=int(m.group(1)); caption=f"{stt}. {items.get(stt,f'Hạng mục {stt}')}"
            if stt not in classified: classified[stt]=[]
            classified[stt].append((path,caption,name_no_ext))
    for stt in classified: classified[stt].sort(key=lambda x:x[2])
    return classified

def wrap_text(draw,text,font,max_w):
    words=text.split(); lines=[]; cur=""
    for w in words:
        test=(cur+" "+w).strip()
        bb=draw.textbbox((0,0),test,font=font)
        if bb[2]-bb[0]<=max_w: cur=test
        else:
            if cur: lines.append(cur)
            cur=w
    if cur: lines.append(cur)
    return lines

def draw_header(page,draw,W,logo_img,bks,gdv):
    draw.rectangle([0,0,W,HEADER_H],fill=NAVY)
    logo_y=(HEADER_H-LOGO_PX)//2
    page.paste(logo_img,(MARGIN,logo_y),logo_img)
    text_x=MARGIN+LOGO_PX+10; text_w=W-text_x-MARGIN
    font_title=get_font(bold=True,pt=16); font_gdv=get_font(bold=False,pt=12)
    title_lines=wrap_text(draw,f"Giám định Giấy tờ xe xe ô tô biển kiểm soát: {bks}",font_title,text_w)
    tlh=pt2px(16)+4; total_th=len(title_lines)*tlh
    gdv_h=pt2px(12)+4 if gdv else 0
    cur_y=(HEADER_H-total_th-gdv_h)//2
    for line in title_lines:
        bb=draw.textbbox((0,0),line,font=font_title)
        x=text_x+(text_w-(bb[2]-bb[0]))//2
        draw.text((x,cur_y),line,font=font_title,fill=WHITE); cur_y+=tlh
    if gdv:
        gdv_text=f"Giám định viên: {gdv}"
        bb=draw.textbbox((0,0),gdv_text,font=font_gdv)
        x=text_x+(text_w-(bb[2]-bb[0]))//2
        draw.text((x,cur_y),gdv_text,font=font_gdv,fill=WHITE)

def draw_footer(page,draw,W,H,page_num,total_pages):
    fy=H-FOOTER_H; draw.rectangle([0,fy,W,H],fill=WHITE)
    draw.line([0,fy,W,fy],fill=GRAY,width=1)
    font_f=get_font(bold=False,pt=10); font_fi=get_font(italic=True,pt=10)
    ty=fy+(FOOTER_H-pt2px(10))//2
    draw.text((MARGIN,ty),"Phòng Giám định và Cứu hộ tại Quảng Ninh",font=font_f,fill=(128,128,128))
    pt=f"Trang {page_num}/{total_pages}"
    bb=draw.textbbox((0,0),pt,font=font_fi)
    draw.text(((W-(bb[2]-bb[0]))//2,ty),pt,font=font_fi,fill=BLACK)

def draw_caption(draw,caption,cx,cy,cell_w,cell_h):
    cap_y=cy+cell_h-CAPTION_H
    draw.rectangle([cx,cap_y,cx+cell_w,cy+cell_h],fill=WHITE)
    draw.line([cx,cap_y,cx+cell_w,cap_y],fill=GRAY,width=1)
    font_cap=get_font(bold=False,pt=11)
    bb=draw.textbbox((0,0),caption,font=font_cap)
    draw.text((cx+(cell_w-(bb[2]-bb[0]))//2, cap_y+(CAPTION_H-(bb[3]-bb[1]))//2),
              caption,font=font_cap,fill=BLACK)

def paste_image_in_cell(page,img_path,cx,cy,cell_w,img_area_h):
    try:
        img=Image.open(img_path).convert('RGB'); iw,ih=img.size
        scale=min(cell_w/iw,img_area_h/ih); nw,nh=int(iw*scale),int(ih*scale)
        img_r=img.resize((nw,nh),Image.LANCZOS)
        page.paste(img_r,(cx+(cell_w-nw)//2, cy+(img_area_h-nh)//2))
    except Exception as e: print(f"  Lỗi ảnh {img_path}: {e}")

def create_landscape_page(imgs,page_num,total,hinfo,logo_img):
    W,H=LANDSCAPE_W,LANDSCAPE_H; COLS,ROWS=2,2
    page=Image.new('RGB',(W,H),WHITE); draw=ImageDraw.Draw(page)
    draw_header(page,draw,W,logo_img,hinfo['bks'],hinfo['gdv'])
    draw_footer(page,draw,W,H,page_num,total)
    draw.rectangle([0,0,W-1,H-1],outline=GRAY,width=1)
    cw=(W-2*MARGIN)//COLS; ch=(H-HEADER_H-FOOTER_H)//ROWS; iah=ch-CAPTION_H
    for idx,(ip,cap) in enumerate(imgs[:COLS*ROWS]):
        col,row=idx%COLS,idx//COLS; cx=MARGIN+col*cw; cy=HEADER_H+row*ch
        draw.rectangle([cx,cy,cx+cw,cy+ch],outline=GRAY,width=1)
        paste_image_in_cell(page,ip,cx,cy,cw,iah)
        draw_caption(draw,cap,cx,cy,cw,ch)
    for idx in range(len(imgs),COLS*ROWS):
        col,row=idx%COLS,idx//COLS; cx=MARGIN+col*cw; cy=HEADER_H+row*ch
        draw.rectangle([cx,cy,cx+cw,cy+ch],outline=GRAY,width=1)
    return page

def create_portrait_page_3(imgs,page_num,total,hinfo,logo_img):
    W,H=PORTRAIT_W,PORTRAIT_H; ROWS=3
    page=Image.new('RGB',(W,H),WHITE); draw=ImageDraw.Draw(page)
    draw_header(page,draw,W,logo_img,hinfo['bks'],hinfo['gdv'])
    draw_footer(page,draw,W,H,page_num,total)
    draw.rectangle([0,0,W-1,H-1],outline=GRAY,width=1)
    cw=W-2*MARGIN; ch=(H-HEADER_H-FOOTER_H)//ROWS; iah=ch-CAPTION_H
    for idx,(ip,cap) in enumerate(imgs[:ROWS]):
        cx=MARGIN; cy=HEADER_H+idx*ch
        draw.rectangle([cx,cy,cx+cw,cy+ch],outline=GRAY,width=1)
        paste_image_in_cell(page,ip,cx,cy,cw,iah)
        draw_caption(draw,cap,cx,cy,cw,ch)
    for idx in range(len(imgs),ROWS):
        cx=MARGIN; cy=HEADER_H+idx*ch
        draw.rectangle([cx,cy,cx+cw,cy+ch],outline=GRAY,width=1)
    return page

def create_portrait_page_1(ip,cap,page_num,total,hinfo,logo_img):
    W,H=PORTRAIT_W,PORTRAIT_H
    page=Image.new('RGB',(W,H),WHITE); draw=ImageDraw.Draw(page)
    draw_header(page,draw,W,logo_img,hinfo['bks'],hinfo['gdv'])
    draw_footer(page,draw,W,H,page_num,total)
    draw.rectangle([0,0,W-1,H-1],outline=GRAY,width=1)
    cx=MARGIN; cy=HEADER_H; cw=W-2*MARGIN; ch=H-HEADER_H-FOOTER_H; iah=ch-CAPTION_H
    draw.rectangle([cx,cy,cx+cw,cy+ch],outline=GRAY,width=1)
    paste_image_in_cell(page,ip,cx,cy,cw,iah)
    draw_caption(draw,cap,cx,cy,cw,ch)
    return page

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Đọc Excel: {EXCEL_PATH}")
    bks, items = read_gtx_excel(EXCEL_PATH, BKS_FALLBACK)
    print(f"  Biển số: {bks}")
    print(f"  Hạng mục: {items}")

    image_paths = [os.path.join(IMAGE_DIR, f) for f in os.listdir(IMAGE_DIR)
                   if f.lower().endswith(('.jpg','.jpeg','.png'))
                   and 'logo' not in f.lower()]
    print(f"\nẢnh tìm thấy ({len(image_paths)}):")
    for p in sorted(image_paths): print(f"  {os.path.basename(p)}")

    classified = classify_images(image_paths, items)
    print(f"\nPhân loại theo STT:")
    for stt in sorted(classified): print(f"  STT {stt}: {len(classified[stt])} ảnh")

    logo_img = process_logo(LOGO_PATH, LOGO_PX)
    hinfo = {'bks': bks, 'gdv': GDV}

    pages_info = []
    imgs_12 = []
    for stt in [1, 2]:
        if stt in classified: imgs_12.extend(classified[stt])
    for i in range(0, max(len(imgs_12), 1), 4):
        chunk = imgs_12[i:i+4]
        if chunk: pages_info.append(('landscape', chunk))

    imgs_34 = []
    for stt in [3, 4]:
        if stt in classified: imgs_34.extend(classified[stt])
    for i in range(0, max(len(imgs_34), 1), 3):
        chunk = imgs_34[i:i+3]
        if chunk: pages_info.append(('portrait3', chunk))

    if 5 in classified:
        for item in classified[5]: pages_info.append(('portrait1', [item]))

    total = len(pages_info)
    print(f"\nTổng số trang: {total}")

    from reportlab.pdfgen import canvas as rlc
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader

    output_path = os.path.join(OUTPUT_DIR, f"GiamDinh_GiayToXe_{bks}.pdf")
    c = rlc.Canvas(output_path)
    for p_idx, (ptype, data) in enumerate(pages_info, 1):
        print(f"  Trang {p_idx}/{total}: {ptype} ({len(data)} ảnh)")
        if ptype == 'landscape':
            pg = create_landscape_page([(d[0],d[1]) for d in data], p_idx, total, hinfo, logo_img)
            pw, ph = 29.7*cm, 21*cm
        elif ptype == 'portrait3':
            pg = create_portrait_page_3([(d[0],d[1]) for d in data], p_idx, total, hinfo, logo_img)
            pw, ph = 21*cm, 29.7*cm
        else:
            d = data[0]
            pg = create_portrait_page_1(d[0], d[1], p_idx, total, hinfo, logo_img)
            pw, ph = 21*cm, 29.7*cm
        c.setPageSize((pw, ph))
        buf = io.BytesIO(); pg.save(buf, format='JPEG', quality=92); buf.seek(0)
        c.drawImage(ImageReader(buf), 0, 0, pw, ph); c.showPage()
    c.save()
    print(f"\n✅ Xuất thành công: {output_path}")

if __name__ == '__main__':
    main()
